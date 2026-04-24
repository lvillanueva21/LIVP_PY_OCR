from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExportadorExcel:
    def __init__(self) -> None:
        self.fill_encabezado = PatternFill(fill_type="solid", fgColor="D9E2EC")
        self.fill_titulo = PatternFill(fill_type="solid", fgColor="BCCCDC")
        self.font_bold = Font(bold=True)
        self.align_top = Alignment(vertical="top", wrap_text=True)

    def exportar_documento(
        self,
        resultado,
        ruta_destino: str,
        *,
        resultado_basico=None,
        resultado_pro=None,
        comparacion=None,
    ) -> None:
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"

        self._llenar_resumen_documento(
            ws_resumen,
            resultado,
            resultado_basico=resultado_basico,
            resultado_pro=resultado_pro,
            comparacion=comparacion,
        )

        ws_paginas = wb.create_sheet("Paginas")
        self._llenar_paginas_documento(ws_paginas, resultado, comparacion=comparacion)

        ws_campos = wb.create_sheet("Campos")
        self._llenar_campos_documento(ws_campos, resultado)

        if comparacion is not None and comparacion.comparaciones_paginas:
            ws_comp = wb.create_sheet("Comparacion paginas")
            self._llenar_comparacion_paginas(ws_comp, comparacion)

        self._guardar_libro(wb, ruta_destino)

    def exportar_historial(self, registros: list[dict], ruta_destino: str) -> None:
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"

        self._llenar_historial_resumen(ws_resumen, registros)

        ws_paginas = wb.create_sheet("Paginas")
        self._llenar_historial_paginas(ws_paginas, registros)

        self._guardar_libro(wb, ruta_destino)

    def _llenar_resumen_documento(
        self,
        ws,
        resultado,
        *,
        resultado_basico=None,
        resultado_pro=None,
        comparacion=None,
    ) -> None:
        filas = [
            ("Archivo", resultado.nombre_archivo),
            ("Ruta", resultado.ruta_archivo),
            ("Páginas", resultado.cantidad_paginas),
            ("Modo mostrado", resultado.etiqueta_modo),
            ("Tiempo total (ms)", resultado.tiempo_total_ms),
            ("Diagnóstico", resultado.diagnostico_general),
            ("Estado OCR", resultado.estado_ocr),
            ("Motor OCR", resultado.motor_ocr),
            ("Recomendación", resultado.recomendacion_modo or "-"),
            ("Observaciones", "\n".join(resultado.observaciones_modo) if resultado.observaciones_modo else "-"),
        ]

        if resultado.metricas_documento_modo is not None:
            metrica = resultado.metricas_documento_modo
            filas.extend(
                [
                    ("Score total", metrica.score_total),
                    ("Score campos", metrica.score_campos),
                    ("Score legibilidad", metrica.score_legibilidad),
                    ("Score confianza", metrica.score_confianza),
                    ("Score texto útil", metrica.score_texto_util),
                    ("Score estabilidad", metrica.score_estabilidad),
                    ("Score velocidad", metrica.score_velocidad),
                    ("Confianza OCR promedio", metrica.confianza_ocr_promedio),
                    ("Confianza OCR mediana", metrica.confianza_ocr_mediana),
                    ("Palabras baja confianza", metrica.palabras_baja_confianza_totales),
                    ("Ruido textual promedio", metrica.ruido_textual_promedio),
                    ("Páginas fáciles", metrica.paginas_faciles),
                    ("Páginas medias", metrica.paginas_medias),
                    ("Páginas difíciles", metrica.paginas_dificiles),
                    ("Páginas críticas", metrica.paginas_criticas),
                    ("Páginas revisión sugerida", metrica.paginas_revision_recomendada),
                ]
            )

        if resultado_basico is not None and resultado_basico.metricas_documento_modo is not None:
            filas.append(("Score Básico", resultado_basico.metricas_documento_modo.score_total))

        if resultado_pro is not None and resultado_pro.metricas_documento_modo is not None:
            filas.append(("Score Pro", resultado_pro.metricas_documento_modo.score_total))

        if comparacion is not None:
            filas.extend(
                [
                    ("Ganador", comparacion.etiqueta_ganador),
                    ("Motivo ganador", comparacion.motivo),
                    ("Revisión manual recomendada", "Sí" if comparacion.revision_manual_recomendada else "No"),
                    ("Motivo revisión manual", comparacion.motivo_revision_manual or "-"),
                ]
            )

        ws["A1"] = "Resumen del documento"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].fill = self.fill_titulo

        fila = 3
        for etiqueta, valor in filas:
            ws.cell(row=fila, column=1, value=etiqueta)
            ws.cell(row=fila, column=2, value=valor)
            ws.cell(row=fila, column=1).font = self.font_bold
            ws.cell(row=fila, column=1).fill = self.fill_encabezado
            ws.cell(row=fila, column=2).alignment = self.align_top
            fila += 1

        self._ajustar_columnas(ws)

    def _llenar_paginas_documento(self, ws, resultado, *, comparacion=None) -> None:
        encabezados = [
            "Página",
            "Fuente",
            "Dificultad",
            "Índice dificultad",
            "Requiere revisión",
            "Confianza promedio",
            "Confianza mediana",
            "Palabras",
            "Palabras baja confianza",
            "Caracteres totales",
            "Tiempo total ms",
            "Tiempo OCR ms",
            "Intentos",
            "Variante",
            "Observaciones",
        ]

        comparacion_mapa = {}
        if comparacion is not None:
            comparacion_mapa = {item.numero_pagina: item for item in comparacion.comparaciones_paginas}

        self._escribir_encabezados(ws, encabezados)

        fila = 2
        for pagina in resultado.resumen_paginas:
            fuente = "texto_digital" if pagina.tiene_texto and pagina.texto_extraido.strip() else "texto_ocr"
            if not pagina.texto_extraido.strip() and not pagina.texto_ocr.strip():
                fuente = "sin_texto"

            comp = comparacion_mapa.get(pagina.numero_pagina)

            valores = [
                pagina.numero_pagina,
                fuente,
                pagina.ocr_dificultad or "-",
                pagina.ocr_dificultad_indice,
                "Sí" if pagina.ocr_requiere_revision else "No",
                pagina.ocr_confianza_promedio,
                pagina.ocr_confianza_mediana,
                pagina.ocr_cantidad_palabras,
                pagina.ocr_palabras_baja_confianza,
                pagina.ocr_caracteres_totales,
                pagina.ocr_tiempo_total_ms,
                pagina.ocr_tiempo_ocr_ms,
                pagina.ocr_numero_intentos,
                pagina.ocr_variante_ganadora or "-",
                "\n".join(pagina.ocr_observaciones) if pagina.ocr_observaciones else "-",
            ]

            if comp is not None:
                if ws.max_column < 19:
                    extras = [
                        "Score Básico",
                        "Score Pro",
                        "Ganador",
                        "Revisión manual comparación",
                    ]
                    for idx, extra in enumerate(extras, start=16):
                        ws.cell(row=1, column=idx, value=extra)
                    self._estilizar_fila_encabezado(ws, 1)
                valores.extend(
                    [
                        comp.score_basico,
                        comp.score_pro,
                        comp.etiqueta_ganador,
                        "Sí" if comp.revision_manual_recomendada else "No",
                    ]
                )

            for columna, valor in enumerate(valores, start=1):
                ws.cell(row=fila, column=columna, value=valor)
                ws.cell(row=fila, column=columna).alignment = self.align_top

            fila += 1

        self._ajustar_columnas(ws)

    def _llenar_campos_documento(self, ws, resultado) -> None:
        encabezados = ["Campo", "Valor", "Detectado", "Estrategia"]
        self._escribir_encabezados(ws, encabezados)

        fila = 2
        for campo in resultado.campos_extraidos:
            valores = [
                campo.etiqueta,
                campo.valor or "No detectado",
                "Sí" if campo.detectado else "No",
                campo.estrategia,
            ]
            for columna, valor in enumerate(valores, start=1):
                ws.cell(row=fila, column=columna, value=valor)
                ws.cell(row=fila, column=columna).alignment = self.align_top
            fila += 1

        self._ajustar_columnas(ws)

    def _llenar_comparacion_paginas(self, ws, comparacion) -> None:
        encabezados = [
            "Página",
            "Score Básico",
            "Score Pro",
            "Ganador",
            "Motivo",
            "Revisión manual",
            "Dificultad Básico",
            "Dificultad Pro",
            "Observaciones",
        ]
        self._escribir_encabezados(ws, encabezados)

        fila = 2
        for item in comparacion.comparaciones_paginas:
            valores = [
                item.numero_pagina,
                item.score_basico,
                item.score_pro,
                item.etiqueta_ganador,
                item.motivo,
                "Sí" if item.revision_manual_recomendada else "No",
                item.dificultad_basico,
                item.dificultad_pro,
                "\n".join(item.observaciones) if item.observaciones else "-",
            ]
            for columna, valor in enumerate(valores, start=1):
                ws.cell(row=fila, column=columna, value=valor)
                ws.cell(row=fila, column=columna).alignment = self.align_top
            fila += 1

        self._ajustar_columnas(ws)

    def _llenar_historial_resumen(self, ws, registros: list[dict]) -> None:
        encabezados = [
            "Fecha y hora",
            "Archivo",
            "Tamaño bytes",
            "Páginas",
            "Modo usado",
            "Score Básico",
            "Score Pro",
            "Ganador",
            "Campos detectados",
            "Tiempo total ms",
            "Revisión manual",
            "Observaciones",
        ]
        self._escribir_encabezados(ws, encabezados)

        fila = 2
        for registro in registros:
            valores = [
                registro.get("fecha_hora"),
                registro.get("archivo_analizado"),
                registro.get("tamano_archivo_bytes"),
                registro.get("cantidad_paginas"),
                registro.get("modo_usado"),
                registro.get("score_basico"),
                registro.get("score_pro"),
                registro.get("ganador"),
                registro.get("campos_detectados"),
                registro.get("tiempo_total_ms"),
                "Sí" if registro.get("revision_manual_recomendada") else "No",
                "\n".join(registro.get("observaciones", [])) or "-",
            ]
            for columna, valor in enumerate(valores, start=1):
                ws.cell(row=fila, column=columna, value=valor)
                ws.cell(row=fila, column=columna).alignment = self.align_top
            fila += 1

        self._ajustar_columnas(ws)

    def _llenar_historial_paginas(self, ws, registros: list[dict]) -> None:
        encabezados = [
            "Fecha y hora",
            "Archivo",
            "Página",
            "Dificultad",
            "Índice dificultad",
            "Requiere revisión",
            "Confianza promedio",
            "Confianza mediana",
            "Palabras",
            "Palabras baja confianza",
            "Tiempo total ms",
            "Intentos",
            "Variante",
            "Observaciones",
        ]
        self._escribir_encabezados(ws, encabezados)

        fila = 2
        for registro in registros:
            for pagina in registro.get("paginas", []):
                valores = [
                    registro.get("fecha_hora"),
                    registro.get("archivo_analizado"),
                    pagina.get("numero_pagina"),
                    pagina.get("dificultad"),
                    pagina.get("indice_dificultad"),
                    "Sí" if pagina.get("requiere_revision") else "No",
                    pagina.get("confianza_promedio"),
                    pagina.get("confianza_mediana"),
                    pagina.get("palabras"),
                    pagina.get("palabras_baja_confianza"),
                    pagina.get("tiempo_total_ms"),
                    pagina.get("intentos"),
                    pagina.get("variante"),
                    "\n".join(pagina.get("observaciones", [])) or "-",
                ]
                for columna, valor in enumerate(valores, start=1):
                    ws.cell(row=fila, column=columna, value=valor)
                    ws.cell(row=fila, column=columna).alignment = self.align_top
                fila += 1

        self._ajustar_columnas(ws)

    def _escribir_encabezados(self, ws, encabezados: list[str]) -> None:
        for columna, encabezado in enumerate(encabezados, start=1):
            ws.cell(row=1, column=columna, value=encabezado)
        self._estilizar_fila_encabezado(ws, 1)

    def _estilizar_fila_encabezado(self, ws, fila: int) -> None:
        for celda in ws[fila]:
            celda.font = self.font_bold
            celda.fill = self.fill_encabezado
            celda.alignment = self.align_top

    def _ajustar_columnas(self, ws) -> None:
        for columna in range(1, ws.max_column + 1):
            letra = get_column_letter(columna)
            max_len = 0
            for fila in range(1, ws.max_row + 1):
                valor = ws.cell(row=fila, column=columna).value
                if valor is None:
                    continue
                texto = str(valor)
                if len(texto) > max_len:
                    max_len = len(texto)
            ws.column_dimensions[letra].width = min(max(14, max_len + 2), 48)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _guardar_libro(self, wb: Workbook, ruta_destino: str) -> None:
        ruta = Path(ruta_destino)
        ruta.parent.mkdir(parents=True, exist_ok=True)
        wb.save(ruta)